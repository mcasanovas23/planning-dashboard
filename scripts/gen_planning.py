import openpyxl, json, sys, os, math
from datetime import datetime, date, timedelta

# ─── Calendario laboral iVascular 2026 ────────────────────────────────────
FACTORY_HOLIDAYS = {
    # Sem 14 – Setmana Santa (cierre completo)
    date(2026,3,27), date(2026,3,30), date(2026,3,31), date(2026,4,1), date(2026,4,2),
    # Sem 15 – Divendres Sant
    date(2026,4,3),
    # Sem 19 – 1 Mayo
    date(2026,5,1),
    # Sem 26 – Sant Joan bridge (solo trabaja 19/06 y 25/06)
    date(2026,6,22), date(2026,6,23), date(2026,6,24),
    # Sem 32-34 – Vacaciones verano
    date(2026,7,31),
    date(2026,8,3),  date(2026,8,4),  date(2026,8,5),  date(2026,8,6),
    date(2026,8,7),  date(2026,8,10), date(2026,8,11), date(2026,8,12), date(2026,8,13),
    date(2026,8,14), date(2026,8,17), date(2026,8,18), date(2026,8,19), date(2026,8,20),
    # Sem 35 – aún vacaciones el viernes 21/08
    date(2026,8,21),
    # Sem 38 – Diada Catalunya
    date(2026,9,11),
    # Sem 42 – Hispanidad
    date(2026,10,12),
    # Sem 50 – puente Constitución/Inmaculada
    date(2026,12,7), date(2026,12,8),
    # Sem 52 – Nochebuena
    date(2026,12,24),
}

def is_working_day(d):
    return d.weekday() < 5 and d not in FACTORY_HOLIDAYS

def add_working_days(start, n):
    """Añade n días laborables a start (n=0 devuelve start si es laboral, sino siguiente)."""
    current = start
    # Si n=0, aseguramos que start sea día laborable
    if n == 0:
        while not is_working_day(current):
            current += timedelta(days=1)
        return current
    count = 0
    while count < n:
        current += timedelta(days=1)
        if is_working_day(current):
            count += 1
    return current

TODAY = date.today()

# Jueves de la semana 23 de fábrica (referencia ancla)
WEEK23_THURSDAY = date(2026, 6, 4)

def thursday_of_week(week_num):
    """Jueves que cierra la semana de fábrica N (semanas empiezan viernes)."""
    try:
        return WEEK23_THURSDAY + timedelta(weeks=(int(week_num) - 23))
    except (TypeError, ValueError):
        return None

EXCEL_PATH = r'C:\Users\mcasanovas\OneDrive - IVASCULAR, S.L.U\Planning General.xlsm'
TMP_PATH   = r'C:\Users\mcasanovas\AppData\Local\Temp\Planning_General_tmp.xlsm'
OUT_HTML        = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'index.html')
OUT_HTML_DADES  = r'P:\DPTO OPERACIONES\04 PLANIFICACION\04 Assistant Planning\Herramientas Web\Dades\Planning General.html'

import shutil
shutil.copy2(EXCEL_PATH, TMP_PATH)

wb = openpyxl.load_workbook(TMP_PATH, read_only=True, keep_vba=False, data_only=True)

# ─── Family merge rules ────────────────────────────────────────────────────
# Families merged into ANGIOLIT (tab), hidden from dashboard as own rows
MERGE_INTO_ANGIOLIT = {'ANGIOBTK', 'PUTNIY', 'ENVIROS'}
# Families merged into ICOVER (tab), hidden from dashboard as own rows
MERGE_INTO_ICOVER   = {'TUVA'}

def canonical_tab(familia):
    if familia in MERGE_INTO_ANGIOLIT: return 'ANGIOLIT'
    if familia in MERGE_INTO_ICOVER:   return 'ICOVER'
    return familia

# ─── Datos ─────────────────────────────────────────────────────────────────
ws = wb['Datos']
all_rows = list(ws.iter_rows(values_only=True))

# Columns: a,b,c,d,f,h,k,m,o,s,t,u,v,w,ab
COL_INDICES = [0,1,2,3,5,7,10,12,14,18,19,21,22,27]
COL_KEYS    = ['article','medidas','familia','lot','setmana',
               'mat1','disp1','mat2','disp2','prep_linia',
               'pzas_lot','sota_cmd','fase','fecha_fab']
COL_HEADERS = ['Artículo','Medidas','Familia','Lote','Semana',
               'Balón','Disp. Mat1','Material 2','Disp. Mat2','Prep. Línea',
               'peces_lot','Sota Cmd','Fase Actual','Fecha Fab.']

rows_data = []
for row in all_rows[1:]:
    if not any(v is not None for v in row[:5]):
        continue
    rd = {}
    for i, ci in enumerate(COL_INDICES):
        v = row[ci] if ci < len(row) else None
        if isinstance(v, datetime): v = v.strftime('%d/%m/%Y')
        rd[COL_KEYS[i]] = v if v is not None else ''
    rd['pzas_cmd'] = row[20] if len(row) > 20 and row[20] is not None else ''
    rd['_retras']  = bool(row[24] == '!!!') if len(row) > 24 else False
    rd['_tab']     = canonical_tab(rd['familia'])  # which tab this row belongs to
    rows_data.append(rd)

# Families that show mat2 columns (explicit list per user spec)
FAM_SHOW_MAT2 = {'ANGIOLIT', 'NAVISCOR', 'ICOVER', 'RESTORER'}
fam_has_mat2 = FAM_SHOW_MAT2

# Build rows grouped by _tab, sorted by setmana then article
rows_by_tab = {}
for rd in rows_data:
    tab = rd['_tab']
    if tab not in rows_by_tab:
        rows_by_tab[tab] = []
    rows_by_tab[tab].append(rd)

for tab in rows_by_tab:
    rows_by_tab[tab].sort(key=lambda r: (
        r['setmana'] if isinstance(r['setmana'], int) else 999,
        str(r['article'])
    ))

# ─── Resum ─────────────────────────────────────────────────────────────────
ws_resum = wb['Resum']
resum_rows = list(ws_resum.iter_rows(values_only=True))

week_row = resum_rows[6]
WEEKS = []
for ci in [4, 8, 12, 16, 20, 24]:
    wnum = week_row[ci]
    if wnum:
        WEEKS.append({'num': int(wnum), 'ci': ci})

# Families to SKIP in dashboard (merged/absorbed elsewhere)
SKIP_IN_DASH = MERGE_INTO_ANGIOLIT | MERGE_INTO_ICOVER

resum_data = []
family_order_resum = []
for row in resum_rows[8:28]:
    familia = row[2]
    if not familia:
        continue
    if familia in SKIP_IN_DASH:
        continue
    family_order_resum.append(familia)
    weeks_info = []
    for w in WEEKS:
        ci = w['ci']
        sem      = row[ci]   if len(row) > ci   else None
        planning = row[ci+1] if len(row) > ci+1 else None
        dif      = row[ci+2] if len(row) > ci+2 else None
        weeks_info.append({'sem': sem, 'planning': planning, 'dif': dif})
    resum_data.append({'familia': familia, 'weeks': weeks_info})

# Tabs ordered: Resum order, only tabs with data
all_tabs_with_data = set(rows_by_tab.keys())
familias_ordered = [f for f in family_order_resum if f in all_tabs_with_data]

# ─── FASES ─────────────────────────────────────────────────────────────────
ws_fases = wb['FASES']
fases_rows = list(ws_fases.iter_rows(values_only=True))
fases_by_family = {}
seen = {}
for row in fases_rows[1:]:
    if not row[0]: continue
    fam = row[0]; name = row[2]; order = row[3]
    if fam not in fases_by_family:
        fases_by_family[fam] = []; seen[fam] = set()
    if name and name not in seen[fam]:
        fases_by_family[fam].append({'name': str(name), 'order': order or 0})
        seen[fam].add(name)
for f in fases_by_family:
    fases_by_family[f].sort(key=lambda x: x['order'])

# Also make merged families share fases
for src in MERGE_INTO_ANGIOLIT:
    if src in fases_by_family and 'ANGIOLIT' not in fases_by_family:
        fases_by_family['ANGIOLIT'] = fases_by_family[src]
for src in MERGE_INTO_ICOVER:
    if src in fases_by_family and 'ICOVER' not in fases_by_family:
        fases_by_family['ICOVER'] = fases_by_family[src]

# ─── Lead time ─────────────────────────────────────────────────────────────
ws_lt = wb['lead time']
leadtime_data = {}
for _lt_row in ws_lt.iter_rows(min_row=2, values_only=True):
    if _lt_row[0] and _lt_row[1]:
        leadtime_data[str(_lt_row[0])] = int(_lt_row[1])

def compute_fecha_fab(familia, tab, fase_actual, fases_map, lt_data):
    if not fase_actual:
        return ''
    phases = fases_map.get(familia) or fases_map.get(tab) or []
    if not phases:
        return ''
    phases_sorted = sorted(phases, key=lambda p: p['order'])
    etiq_order = next(
        (p['order'] for p in phases_sorted if 'etiquetado envase primario' in p['name'].lower()),
        None
    )
    if not etiq_order:
        return ''
    lt = lt_data.get(familia) or lt_data.get(tab)
    if not lt:
        return ''
    phases_per_day = etiq_order / lt
    fase_lower = str(fase_actual).lower().strip()
    current_order = next(
        (p['order'] for p in phases_sorted
         if fase_lower[:15] in p['name'].lower() or p['name'].lower()[:15] in fase_lower),
        None
    )
    if current_order is None:
        return ''
    if current_order >= etiq_order:
        return '__ESTERIL__'
    remaining_phases = etiq_order - current_order
    remaining_days = math.ceil(remaining_phases / phases_per_day)
    fecha = add_working_days(TODAY, remaining_days - 1)
    return fecha.strftime('%d/%m/%Y')

# Second pass: add fecha_fab and fecha_obj to each row
for rd in rows_data:
    rd['fecha_fab'] = compute_fecha_fab(rd['familia'], rd['_tab'], rd.get('fase', ''), fases_by_family, leadtime_data)
    d = thursday_of_week(rd.get('setmana', ''))
    rd['fecha_obj'] = d.strftime('%d/%m/%Y') if d else ''

# ─── JSON blobs ─────────────────────────────────────────────────────────────
j_resum    = json.dumps(resum_data,      ensure_ascii=False)
j_weeks    = json.dumps([w['num'] for w in WEEKS], ensure_ascii=False)
j_fam_ord  = json.dumps(familias_ordered, ensure_ascii=False)
j_col_keys = json.dumps(COL_KEYS,        ensure_ascii=False)
j_col_hdr  = json.dumps(COL_HEADERS,     ensure_ascii=False)
j_fases    = json.dumps(fases_by_family, ensure_ascii=False)
j_all_rows = json.dumps(rows_by_tab,     ensure_ascii=False)
j_has_mat2 = json.dumps(list(fam_has_mat2), ensure_ascii=False)

ts = datetime.now().strftime('%d/%m/%Y %H:%M')

# ─── HTML ───────────────────────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Planning General – iVascular (CATBAL)</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#1e293b;font-size:13px}}

/* ── Top bar ── */
.top-bar{{background:#0f2044;color:#fff;padding:10px 20px;display:flex;align-items:center;gap:16px;
  position:sticky;top:0;z-index:200;box-shadow:0 2px 8px rgba(0,0,0,.5)}}
.top-bar h1{{font-size:15px;font-weight:700;letter-spacing:.5px;white-space:nowrap}}
.top-bar .ts{{font-size:10px;color:#64748b;margin-left:auto;white-space:nowrap}}

/* ── Tab nav ── */
.tab-nav{{background:#162d5a;display:flex;flex-wrap:wrap;gap:2px;padding:0 8px;
  position:sticky;top:41px;z-index:190;box-shadow:0 2px 4px rgba(0,0,0,.3)}}
.tab-btn{{padding:7px 13px;border:none;background:transparent;color:#94a3b8;cursor:pointer;
  font-size:11px;font-weight:600;border-bottom:3px solid transparent;transition:all .15s;white-space:nowrap}}
.tab-btn:hover{{color:#e2e8f0;background:rgba(255,255,255,.06)}}
.tab-btn.active{{color:#fff;border-bottom-color:#3b82f6;background:rgba(255,255,255,.08)}}

/* ── Content ── */
.tab-content{{display:none;padding:20px}}
.tab-content.active{{display:block}}

/* ── Dashboard ── */
.dash-header{{display:flex;align-items:baseline;gap:12px;margin-bottom:14px}}
.dash-header h2{{font-size:17px;font-weight:700;color:#0f2044}}
.dash-meta{{font-size:11px;color:#64748b}}

.table-wrap{{overflow-x:auto;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,.1)}}
.plan-table-wrap{{overflow:auto;max-height:calc(100vh - 200px)}}

/* ── Dashboard table — responsive ── */
.dash-table{{width:100%;table-layout:fixed;border-collapse:collapse;background:#fff;font-size:12px}}
.dash-table thead tr.wk-header th{{
  background:#0f2044;color:#fff;padding:9px 8px;text-align:center;
  font-size:11px;font-weight:700;letter-spacing:.4px;overflow:hidden;text-overflow:ellipsis;
  border-right:1px solid rgba(255,255,255,.1)}}
.dash-table thead tr.wk-header th:first-child{{text-align:left;border-right:2px solid rgba(255,255,255,.2)}}
.dash-table thead tr.wk-header th:last-child{{border-right:none}}
.dash-table tbody tr{{border-bottom:1px solid #f1f5f9;transition:background .1s}}
.dash-table tbody tr:hover{{background:#f0f7ff}}
.dash-table td{{padding:8px 8px;vertical-align:middle;border-right:1px solid #f1f5f9;font-size:11px;overflow:hidden}}
.dash-table td:first-child{{font-weight:700;color:#0f2044;font-size:12px;border-right:2px solid #e2e8f0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;text-align:center}}
.dash-table td:last-child{{border-right:none}}
.dash-table thead tr.wk-subheader th{{background:#0a1a38;color:#94a3b8;padding:4px 8px;
  text-align:center;font-size:10px;font-weight:500;border-right:1px solid rgba(255,255,255,.08)}}
.dash-table thead tr.wk-subheader th.prog-sub{{color:#86efac}}
.dash-table td.week-pair-start{{border-left:2px solid #e2e8f0}}
.dash-table tbody tr:last-child td{{border-bottom:none}}

/* Week cell: planning (bold) / sem (grey small)  chip — todo en línea */
.week-cell{{display:flex;align-items:center;gap:5px;white-space:nowrap;justify-content:center}}
.wc-planning{{font-weight:700;font-size:12px;color:#1e293b}}
.wc-sep{{color:#cbd5e1;font-size:11px;margin:0 1px}}
.wc-sem{{font-size:10px;color:#94a3b8;font-weight:400}}
.dif-chip{{display:inline-block;padding:1px 6px;border-radius:8px;font-size:10px;font-weight:700;white-space:nowrap}}
.dif-chip.pos{{background:#dcfce7;color:#15803d}}
.dif-chip.warn{{background:#fef9c3;color:#854d0e}}
.dif-chip.neg{{background:#fee2e2;color:#b91c1c}}
.no-data-cell{{color:#d1d5db;font-size:11px;text-align:center}}

.detail-btn{{padding:4px 10px;background:#0f2044;color:#fff;border:none;border-radius:5px;
  font-size:11px;font-weight:600;cursor:pointer;transition:background .15s;margin:0 6px}}
.detail-btn:hover{{background:#1e3a6e}}
.dash-table tbody tr.totals-row td{{font-weight:700;background:#f8fafc;border-top:2px solid #e2e8f0;color:#0f2044}}

/* ── Colored tabs (fucsia/rosa) ── */
.tab-btn.tab-rosa{{color:#f9a8d4}}
.tab-btn.tab-rosa:hover{{color:#f472b6;background:rgba(244,114,182,.08)}}
.tab-btn.tab-rosa.active{{color:#f72585;border-bottom-color:#f72585;background:rgba(247,37,133,.08)}}

/* ── Family tabs ── */
.fam-controls{{display:flex;align-items:center;gap:10px;margin-bottom:14px;flex-wrap:wrap}}
.fam-controls h2{{font-size:17px;font-weight:700;color:#0f2044}}
.search-box{{padding:5px 11px;border:1px solid #cbd5e1;border-radius:6px;font-size:12px;
  width:190px;outline:none}}
.search-box:focus{{border-color:#3b82f6;box-shadow:0 0 0 2px rgba(59,130,246,.2)}}
.row-count{{font-size:11px;color:#64748b;margin-left:auto}}

/* ── Planning table ── */
.plan-table{{width:100%;border-collapse:collapse;background:#fff;font-size:12px}}
.plan-table thead tr:first-child th{{background:#0f2044;color:#fff;padding:8px 10px;text-align:center;
  white-space:nowrap;position:sticky;top:0;z-index:11;font-size:11px;font-weight:700;
  letter-spacing:.3px}}
.plan-table tbody tr{{border-bottom:1px solid #f1f5f9;transition:background .1s}}
.plan-table tbody tr:hover{{background:#f0f7ff}}
/* retras: sin fondo rojo, solo hover normal */
.plan-table td{{padding:6px 10px;vertical-align:middle;white-space:nowrap;text-align:center}}

/* ── Cell styles ── */
.badge{{display:inline-block;padding:1px 7px;border-radius:10px;font-size:10px;font-weight:700;white-space:nowrap}}
.badge-asig  {{background:#dcfce7;color:#166534}}
.badge-stock {{background:#dbeafe;color:#1e40af}}
.badge-proc  {{background:#fef9c3;color:#854d0e}}
.badge-pdt   {{background:#fee2e2;color:#991b1b}}
.badge-pdtmq {{background:#fce7f3;color:#9d174d}}
.badge-txt   {{background:#f3f4f6;color:#374151;border:1px solid #e5e7eb}}
.badge-prio  {{background:#f97316;color:#fff}}

/* Sota comanda: badge rojo */
.sota-check{{
  display:inline-block;padding:2px 8px;border-radius:10px;
  background:#d32f2f;color:#fff;font-size:10px;font-weight:700;
  white-space:nowrap}}
/* Prep línea: verde Asignado */
.prep-check{{
  display:inline-flex;align-items:center;justify-content:center;
  width:15px;height:15px;border-radius:3px;
  background:#dcfce7;color:#166534;font-size:10px;font-weight:700;
  line-height:1;border:1px solid #bbf7d0}}

/* Fase progress */
.fase-wrap{{display:flex;flex-direction:column;gap:2px}}
.fase-text{{font-size:11px;color:#374151;max-width:190px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.fase-bar-bg{{height:3px;background:#e5e7eb;border-radius:2px;width:100px}}
.fase-bar{{height:3px;background:#3b82f6;border-radius:2px}}

/* ── Column filters ── */
.filter-row th{{background:#162d5a;padding:3px 4px;position:sticky;top:34px;z-index:9}}
.col-filter-wrap{{position:relative;text-align:center}}
.col-filter-btn{{width:100%;padding:2px 4px;border:1px solid rgba(255,255,255,.2);border-radius:4px;
  background:rgba(255,255,255,.08);color:#e2e8f0;font-size:10px;cursor:pointer;font-family:inherit}}
.col-filter-btn:hover{{background:rgba(255,255,255,.18)}}
.col-filter-btn.active{{background:rgba(59,130,246,.35);border-color:#60a5fa;color:#fff}}
.filter-dropdown{{background:#fff;border:1px solid #e2e8f0;border-radius:8px;
  box-shadow:0 4px 16px rgba(0,0,0,.18);min-width:160px;max-width:260px;padding:6px 0;z-index:9999}}
.fd-search{{display:block;width:calc(100% - 16px);margin:0 8px 4px;padding:4px 8px;
  border:1px solid #cbd5e1;border-radius:5px;font-size:11px;outline:none}}
.fd-search:focus{{border-color:#3b82f6}}
.fd-options{{max-height:220px;overflow-y:auto}}
.fd-option{{display:flex;align-items:center;gap:7px;padding:4px 12px;cursor:pointer;
  font-size:11px;color:#374151;white-space:nowrap}}
.fd-option:hover{{background:#f0f7ff}}
.fd-option input[type=checkbox]{{cursor:pointer;flex-shrink:0;accent-color:#3b82f6}}
.fd-all{{border-bottom:1px solid #f1f5f9;margin-bottom:2px;padding-bottom:5px;color:#64748b}}
/* Esterilizando badge */
.esteril-badge{{display:inline-block;padding:2px 8px;border-radius:10px;
  background:#0f2044;color:#fff;font-size:10px;font-weight:700;white-space:nowrap}}

/* Scrollbar */
::-webkit-scrollbar{{height:6px;width:6px}}
::-webkit-scrollbar-track{{background:#f1f5f9}}
::-webkit-scrollbar-thumb{{background:#cbd5e1;border-radius:3px}}
</style>
</head>
<body>

<div class="top-bar">
  <h1>&#9783; Planning General – iVascular (CATBAL) <span style="font-size:11px;font-weight:400;color:#94a3b8;margin-left:8px">{ts}</span></h1>
</div>

<nav class="tab-nav" id="tab-nav">
  <button class="tab-btn active" data-id="dashboard" onclick="showTab('dashboard')">Dashboard</button>
</nav>

<div id="dashboard" class="tab-content active">
  <div class="dash-header">
    <h2>Resumen semanal</h2>
    <span class="dash-meta" id="dash-meta"></span>
  </div>
  <div class="table-wrap">
    <table class="dash-table" id="dash-table"></table>
  </div>
</div>

<script>
const RESUM_DATA   = {j_resum};
const WEEKS        = {j_weeks};
const FAM_ORDERED  = {j_fam_ord};
const COL_KEYS     = {j_col_keys};
const COL_HEADERS  = {j_col_hdr};
const FASES_MAP    = {j_fases};
const ALL_ROWS     = {j_all_rows};
const FAM_HAS_MAT2 = new Set({j_has_mat2});

// ── Tabs ──────────────────────────────────────────────────────────────────
const searchState    = {{}};
const colFilterState = {{}};

function showTab(id) {{
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
  if (!document.getElementById(id)) buildFamiliaTab(id);
  document.getElementById(id).classList.add('active');
  document.querySelectorAll('.tab-btn').forEach(btn => {{
    if (btn.dataset.id === id) btn.classList.add('active');
  }});
}}

const ROSA_TABS = new Set(['ANGIOLIT','NCXPERIE','XP PRO','ESS PRO','NAVISCOR']);

function buildNav() {{
  const nav = document.getElementById('tab-nav');
  FAM_ORDERED.forEach(fam => {{
    const btn = document.createElement('button');
    btn.className = 'tab-btn' + (ROSA_TABS.has(fam) ? ' tab-rosa' : '');
    btn.dataset.id = fam;
    btn.textContent = fam;
    btn.onclick = () => showTab(fam);
    nav.appendChild(btn);
  }});
}}

// ── Dashboard ─────────────────────────────────────────────────────────────
function famWeekProgress(fam, weekNum) {{
  const rows = (ALL_ROWS[fam] || []).filter(r => r.setmana == weekNum);
  if (!rows.length) return null;
  const phases = FASES_MAP[fam] || [];
  const total  = phases.length;
  if (!total) return null;
  let sum = 0, count = 0;
  rows.forEach(r => {{
    if (!r.fase) return;
    const idx = phases.findIndex(p => r.fase.startsWith(p.name.substring(0, 12)));
    if (idx >= 0) {{ sum += Math.round(((idx + 1) / total) * 100); count++; }}
  }});
  return count ? Math.round(sum / count) : 0;
}}

function difChip(val) {{
  if (val === null || val === undefined) return '<span class="no-data-cell">—</span>';
  const n = Number(val);
  if (n > 15)  return `<span class="dif-chip neg">${{n}}</span>`;
  if (n < -15) return `<span class="dif-chip warn">${{n}}</span>`;
  return `<span class="dif-chip pos">${{n}}</span>`;
}}

function buildDashboard() {{
  const tbl = document.getElementById('dash-table');
  const nWeeks = WEEKS.length;

  // Colgroup: familia fija + por semana: planif | teórico | dif | progreso
  let cg = `<colgroup><col style="width:110px">`;
  for (let i = 0; i < nWeeks; i++) cg += `<col style="width:44px"><col style="width:44px"><col style="width:52px"><col style="width:85px">`;
  cg += `</colgroup>`;

  // Header: fila superior colspan=4 por semana, fila inferior con 4 subetiquetas
  let th1 = `<th rowspan="2" style="vertical-align:middle">Familia</th>`;
  let th2 = ``;
  WEEKS.forEach(w => {{
    th1 += `<th colspan="4" style="border-left:2px solid rgba(255,255,255,.25);border-bottom:1px solid rgba(255,255,255,.12)">Sem. ${{w}}</th>`;
    th2 += `<th>Teór.</th><th>Plan.</th><th>Dif.</th><th class="prog-sub">Prog.</th>`;
  }});
  tbl.innerHTML = cg + `<thead>
    <tr class="wk-header">${{th1}}</tr>
    <tr class="wk-subheader">${{th2}}</tr>
  </thead><tbody id="dash-tbody"></tbody>`;

  const tbody = document.getElementById('dash-tbody');
  RESUM_DATA.forEach(fam => {{
    let cells = fam.weeks.map((w, wi) => {{
      const weekNum  = WEEKS[wi];
      const pct      = famWeekProgress(fam.familia, weekNum);
      const progCell = pct !== null
        ? `<td><div style="display:flex;align-items:center;gap:4px;padding:0 6px">
            <div style="flex:1;height:5px;background:#e5e7eb;border-radius:3px;min-width:20px;overflow:hidden">
              <div style="height:5px;background:#22c55e;border-radius:3px;width:${{pct}}%"></div>
            </div>
            <span style="font-size:10px;font-weight:700;color:#1e293b;white-space:nowrap">${{pct}}%</span>
          </div></td>`
        : `<td></td>`;
      if (w.sem === null && w.planning === null)
        return `<td class="week-pair-start"></td><td></td><td></td>${{progCell}}`;
      return `<td class="week-pair-start" style="text-align:center;font-size:10px;color:#94a3b8">${{w.sem ?? '—'}}</td>` +
             `<td style="text-align:center;font-weight:700;font-size:12px;color:#1e293b">${{w.planning ?? '—'}}</td>` +
             `<td style="text-align:center">${{difChip(w.dif)}}</td>` +
             progCell;
    }}).join('');

    tbody.insertAdjacentHTML('beforeend',
      `<tr><td>${{fam.familia}}</td>${{cells}}</tr>`);
  }});

  // Totals row
  const totals = WEEKS.map((w, wi) => {{
    let sumP = 0, sumS = 0;
    RESUM_DATA.forEach(fam => {{
      const wk = fam.weeks[wi];
      if (wk && wk.planning !== null && wk.planning !== undefined) sumP += Number(wk.planning) || 0;
      if (wk && wk.sem !== null && wk.sem !== undefined) sumS += Number(wk.sem) || 0;
    }});
    return {{planning: sumP, sem: sumS}};
  }});
  const totalCells = totals.map(t =>
    `<td class="week-pair-start" style="text-align:center;font-size:10px;color:#94a3b8">${{t.sem}}</td>` +
    `<td style="text-align:center;font-weight:700;font-size:12px;color:#1e293b">${{t.planning}}</td>` +
    `<td></td><td></td>`
  ).join('');
  tbody.insertAdjacentHTML('beforeend',
    `<tr class="totals-row"><td>Total</td>${{totalCells}}</tr>`);

  document.getElementById('dash-meta').textContent =
    `Semanas ${{WEEKS[0]}} – ${{WEEKS[WEEKS.length-1]}} · ${{RESUM_DATA.length}} familias`;
}}

// ── Family tab ────────────────────────────────────────────────────────────
function dispBadge(val) {{
  if (!val) return '';
  const v = (val+'').toLowerCase();
  if (v === 'asignado')           return `<span class="badge badge-asig">${{val}}</span>`;
  if (v === 'en stock')           return `<span class="badge badge-stock">${{val}}</span>`;
  if (v.includes('proceso'))      return `<span class="badge badge-proc">${{val}}</span>`;
  if (v.includes('pdt. mq'))      return `<span class="badge badge-pdtmq">${{val}}</span>`;
  if (v.startsWith('pdt'))        return `<span class="badge badge-pdt">${{val}}</span>`;
  return `<span class="badge">${{val}}</span>`;
}}

function parseDMY(s) {{
  if (!s || s === '__ESTERIL__') return null;
  const p = s.split('/');
  return p.length === 3 ? new Date(+p[2], +p[1]-1, +p[0]) : null;
}}

function txtBadge(val) {{
  if (!val) return '';
  if ((val+'').toLowerCase().includes('prio')) return `<span class="badge badge-prio">${{val}}</span>`;
  return `<span class="badge badge-txt">${{val}}</span>`;
}}

function faseCell(fam, fase) {{
  if (!fase) return '';
  const phases = FASES_MAP[fam] || [];
  if (!phases.length) return `<div class="fase-wrap"><span class="fase-text" title="${{fase}}">${{fase}}</span></div>`;
  const etiqIdx = phases.findIndex(p => p.name.toLowerCase().includes('etiquetado envase primario'));
  const capIdx  = etiqIdx >= 0 ? etiqIdx : phases.length - 1;
  const idx     = phases.findIndex(p => fase.startsWith(p.name.substring(0, 12)));
  let pct = 0;
  if (idx >= 0) pct = idx >= capIdx ? 100 : Math.round(((idx + 1) / (capIdx + 1)) * 100);
  const barColor = pct === 100 ? '#22c55e' : '#3b82f6';
  const barHtml = idx >= 0
    ? `<div style="display:flex;align-items:center;gap:5px">
        <div class="fase-bar-bg"><div class="fase-bar" style="width:${{pct}}%;background:${{barColor}}"></div></div>
        <span style="font-size:10px;color:#1e293b;font-weight:600">${{pct}}%</span>
      </div>`
    : '';
  return `<div class="fase-wrap">
    <span class="fase-text" title="${{fase}}">${{fase}}</span>
    ${{barHtml}}
  </div>`;
}}

function renderTable(fam) {{
  const rows    = ALL_ROWS[fam] || [];
  const search  = (searchState[fam] || '').toLowerCase();
  const showMat2 = FAM_HAS_MAT2.has(fam);

  const colFilters = colFilterState[fam] || {{}};
  const filtered = rows.filter(r => {{
    if (search && !Object.values(r).join(' ').toLowerCase().includes(search)) return false;
    for (const [k, v] of Object.entries(colFilters)) {{
      if (v === null || v === undefined) continue;
      if (!v.has(String(r[k] ?? ''))) return false;
    }}
    return true;
  }});

  document.getElementById('rowcount-' + fam).textContent =
    `${{filtered.length}} / ${{rows.length}} filas`;

  const tbody = document.getElementById('tbody-' + fam);
  tbody.innerHTML = filtered.map(r => {{
    const cells = COL_KEYS.map((k, i) => {{
      // Skip mat2/disp2 columns if family has no mat2 data
      if ((k === 'mat2' || k === 'disp2') && !showMat2) return '';

      const val = r[k];
      if (k === 'disp1' || k === 'disp2')  return `<td>${{dispBadge(val)}}</td>`;
      if (k === 'txt')                      return `<td>${{txtBadge(val)}}</td>`;
      if (k === 'fase')       return `<td style="text-align:left">${{faseCell(r.familia || fam, val)}}</td>`;
      if (k === 'fecha_fab') {{
        if (val === '__ESTERIL__') return `<td><span class="esteril-badge">Esterilizando</span></td>`;
        if (val) {{
          const dFab = parseDMY(val), dObj = parseDMY(r.fecha_obj);
          if (dFab && dObj && dFab > dObj)
            return `<td><span class="sota-check">${{val}}</span></td>`;
        }}
        return `<td style="font-size:11px;color:#374151;white-space:nowrap">${{val ?? ''}}</td>`;
      }}
      if (k === 'article')   return `<td style="font-family:monospace;font-size:11px">${{val}}</td>`;
      if (k === 'sota_cmd')   return `<td>${{val === 'X' ? '<span class="sota-check">comanda</span>' : ''}}</td>`;
      if (k === 'prep_linia') return `<td>${{val === 'X' ? '<span class="prep-check">✓</span>' : ''}}</td>`;
      if (k === 'pzas_lot') return `<td><div style="display:inline-flex;align-items:flex-end;gap:1px"><span style="font-weight:600">${{val ?? ''}}</span>${{r.pzas_cmd ? `<span style="font-size:10px;color:#94a3b8;line-height:1.4">/${{r.pzas_cmd}}</span>` : ''}}</div></td>`;
      if (k === 'setmana')   return `<td style="font-weight:700;color:#0f2044">${{val}}</td>`;
      if (k === 'lot')       return `<td style="font-family:monospace;font-size:11px">${{val}}</td>`;
      return `<td>${{val ?? ''}}</td>`;
    }}).join('');

    return `<tr>${{cells}}</tr>`;
  }}).join('');
}}

function buildFamiliaTab(fam) {{
  const rows    = ALL_ROWS[fam] || [];
  const showMat2 = FAM_HAS_MAT2.has(fam);

  const thCols = COL_KEYS.map((k, i) => {{
    if ((k === 'mat2' || k === 'disp2') && !showMat2) return '';
    return `<th>${{COL_HEADERS[i]}}</th>`;
  }}).join('');

  const filterCells = COL_KEYS.map((k) => {{
    if ((k === 'mat2' || k === 'disp2') && !showMat2) return '';
    const sid = _sid(fam, k);
    return `<th><div class="col-filter-wrap">
      <button class="col-filter-btn" id="cfb_${{sid}}"
        onclick="openColFilter('${{fam.replace(/'/g,"\\'")}}','${{k}}',this)">▼</button>
    </div></th>`;
  }}).join('');

  const div = document.createElement('div');
  div.id = fam; div.className = 'tab-content';
  div.innerHTML = `
    <div class="fam-controls">
      <h2>${{fam}}</h2>
      <input class="search-box" placeholder="Buscar en todo..." oninput="onSearch('${{fam}}',this.value)">
      <span class="row-count" id="rowcount-${{fam}}"></span>
    </div>
    <div class="table-wrap plan-table-wrap">
      <table class="plan-table">
        <thead>
          <tr>${{thCols}}</tr>
          <tr class="filter-row">${{filterCells}}</tr>
        </thead>
        <tbody id="tbody-${{fam}}"></tbody>
      </table>
    </div>`;
  document.body.appendChild(div);
  renderTable(fam);
}}

function onSearch(fam, val) {{ searchState[fam] = val; renderTable(fam); }}

// ── Excel-style column filters ────────────────────────────────────────────
function _sid(fam, key) {{ return (fam + '_' + key).replace(/[ .]/g,'_'); }}

let _fdFam = null, _fdKey = null, _fdClose = null;

function getColVals(fam, key) {{
  return [...new Set((ALL_ROWS[fam]||[]).map(r => String(r[key]??'')))].sort((a,b)=>{{
    const na=Number(a),nb=Number(b); return (!isNaN(na)&&!isNaN(nb))?na-nb:a.localeCompare(b);
  }});
}}

function openColFilter(fam, key, anchor) {{
  document.querySelectorAll('.filter-dropdown').forEach(d=>d.remove());
  if (_fdClose) {{ document.removeEventListener('mousedown',_fdClose); _fdClose=null; }}
  if (_fdFam===fam && _fdKey===key) {{ _fdFam=null; _fdKey=null; return; }}
  _fdFam=fam; _fdKey=key;

  const allVals = getColVals(fam, key);
  const sel = (colFilterState[fam]||{{}})[key]; // null/undef=all, Set=filtered
  const allSel = sel===null||sel===undefined;
  const sid = _sid(fam, key);

  const div = document.createElement('div');
  div.className = 'filter-dropdown';
  div.innerHTML = `<input class="fd-search" placeholder="Buscar..." oninput="fdSearch(this,'${{sid}}')">
    <div class="fd-options" id="fdo_${{sid}}">
      <label class="fd-option fd-all">
        <input type="checkbox" ${{allSel?'checked':''}} onchange="fdAll(this,'${{fam}}','${{key}}')">
        <span><em>(Seleccionar todo)</em></span></label>
      ${{allVals.map(v=>`<label class="fd-option" data-v="${{v.replace(/"/g,'&quot;')}}">
        <input type="checkbox" value="${{v.replace(/"/g,'&quot;')}}" ${{(allSel||sel.has(v))?'checked':''}}
          onchange="fdVal(this,'${{fam}}','${{key}}')">
        <span>${{v===''?'(vacío)':v}}</span></label>`).join('')}}
    </div>`;
  document.body.appendChild(div);
  const r=anchor.getBoundingClientRect();
  div.style.cssText=`position:fixed;top:${{r.bottom+2}}px;left:${{Math.max(0,r.left)}}px`;
  _fdClose=function(e){{
    if(!div.contains(e.target)&&e.target!==anchor){{
      div.remove();_fdFam=null;_fdKey=null;
      document.removeEventListener('mousedown',_fdClose);_fdClose=null;
    }}
  }};
  setTimeout(()=>document.addEventListener('mousedown',_fdClose),50);
}}

function fdSearch(inp, sid) {{
  const q=inp.value.toLowerCase();
  document.querySelectorAll(`#fdo_${{sid}} .fd-option:not(.fd-all)`).forEach(el=>{{
    el.style.display=el.querySelector('span').textContent.toLowerCase().includes(q)?'':'none';
  }});
}}

function fdAll(cb, fam, key) {{
  if(!colFilterState[fam]) colFilterState[fam]={{}};
  const sid=_sid(fam,key);
  if(cb.checked) {{
    colFilterState[fam][key]=null;
    document.querySelectorAll(`#fdo_${{sid}} input[type=checkbox]`).forEach(c=>c.checked=true);
  }} else {{
    colFilterState[fam][key]=new Set();
    document.querySelectorAll(`#fdo_${{sid}} .fd-option:not(.fd-all) input`).forEach(c=>c.checked=false);
  }}
  renderTable(fam); _fdBtnUpdate(fam,key);
}}

function fdVal(cb, fam, key) {{
  if(!colFilterState[fam]) colFilterState[fam]={{}};
  const allVals=getColVals(fam,key);
  let sel=colFilterState[fam][key];
  if(sel===null||sel===undefined) sel=new Set(allVals);
  cb.checked?sel.add(cb.value):sel.delete(cb.value);
  colFilterState[fam][key]=sel.size===allVals.length?null:sel;
  const allCb=document.querySelector(`#fdo_${{_sid(fam,key)}} .fd-all input`);
  if(allCb) allCb.checked=colFilterState[fam][key]===null;
  renderTable(fam); _fdBtnUpdate(fam,key);
}}

function _fdBtnUpdate(fam, key) {{
  const btn=document.getElementById(`cfb_${{_sid(fam,key)}}`);
  if(btn) btn.classList.toggle('active', colFilterState[fam]?.[key]!==null&&colFilterState[fam]?.[key]!==undefined);
}}

// ── Init ──────────────────────────────────────────────────────────────────
buildNav();
buildDashboard();
</script>
</body>
</html>"""

with open(OUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

try:
    os.makedirs(os.path.dirname(OUT_HTML_DADES), exist_ok=True)
    shutil.copy2(OUT_HTML, OUT_HTML_DADES)
    print(f"OK: {OUT_HTML_DADES}")
except Exception as e:
    print(f"AVISO: no se pudo copiar a Dades: {e}")

print(f"OK: {OUT_HTML}")
print(f"Familias dashboard: {len(resum_data)}")
print(f"Tabs con datos: {familias_ordered}")
print(f"Tabs con mat2: {sorted(fam_has_mat2)}")
